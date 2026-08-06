import io

import openpyxl
import pytest

from app.companies.import_xlsx import ParsedRow, XlsxParseError, parse_workbook, site_key


def _make_workbook(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Запрос", "Название", "Категории", "Регион", "Город", "Полный адрес",
              "Мобильные", "Немобильные", "Сайт", "Email с сайта компании", "График",
              "Широта", "Долгота", "Оценок", "Отзывов", "Рейтинг"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_site_key_normalizes_url():
    assert site_key("https://www.Stroyka.ru/") == "stroyka.ru"
    assert site_key("http://stroyka.ru") == "stroyka.ru"
    assert site_key("") == ""


def test_category_raw_takes_first_segment_before_pipe():
    data = _make_workbook([
        ["застройщик", "ООО Дом", "Строительство дачных домов и коттеджей | бани | стройка",
         "Самарская область", "Самара", "ул. Ленина 1", "", "+7 846 000-00-00",
         "https://dom-samara.ru", "info@dom-samara.ru", "", "", "", 10, 5, 4.8],
    ])
    rows = parse_workbook(data)
    assert rows[0].category_raw == "Строительство дачных домов и коттеджей"


def test_row_without_site_is_dropped():
    data = _make_workbook([
        ["застройщик", "ООО Без сайта", "Категория", "Самара", "Самара", "", "", "",
         "", "", "", "", "", 0, 0, None],
    ])
    rows = parse_workbook(data)
    assert rows == []


def test_missing_required_column_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Название"])   # нет "Сайт", "Регион" и т.д.
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(XlsxParseError):
        parse_workbook(buf.getvalue())


def test_duplicate_site_key_within_file_collapses_to_one_row():
    data = _make_workbook([
        ["з", "ООО Дом", "Кат", "Самара", "Самара", "", "", "", "https://dom.ru", "", "",
         "", "", 5, 3, 4.5],
        ["з", "ООО Дом (дубль)", "Кат", "Самара", "Самара", "", "", "",
         "https://www.dom.ru/", "", "", "", "", 8, 6, 4.7],
    ])
    rows = parse_workbook(data)
    assert len(rows) == 1
    assert rows[0].reviews_count == 6   # последняя встреченная строка побеждает


def test_short_row_does_not_crash():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Название", "Категории", "Регион", "Город", "Сайт"])   # короткий заголовок, без числовых колонок
    ws.append(["ООО Дом", "Дома", "Самара", "Самара", "https://dom.ru"])
    buf = io.BytesIO()
    wb.save(buf)
    rows = parse_workbook(buf.getvalue())
    assert len(rows) == 1
    assert rows[0].reviews_count == 0


def test_empty_workbook_raises_parse_error():
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(XlsxParseError):
        parse_workbook(buf.getvalue())
